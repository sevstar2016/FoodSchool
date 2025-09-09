from io import BytesIO
from datetime import date, timedelta
from collections import defaultdict

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import db_session, require_admin
from app.models.users import User, Class
from app.models.complexes import UserComplexChoice, Complex, ComplexWeekday


router = APIRouter(prefix="/exports", tags=["exports"]) 


def _current_monday(today: date) -> date:
    return today - timedelta(days=today.weekday())


def _last_week_monday(today: date) -> date:
    return _current_monday(today) - timedelta(days=7)


def _next_monday(today: date) -> date:
    return today + timedelta(days=((7 - today.weekday()) % 7))


def _resolve_week_start(db: Session, mode: str | None, explicit: date | None) -> date:
    """
    Resolve which week_start to export:
    - if explicit provided -> use it
    - if mode == 'last' -> last completed week's Monday
    - if mode == 'current' -> current week's Monday
    - if mode == 'next' -> next Monday
    - else (None or 'latest') -> latest available week_start in user_complex_choices
    """
    today = date.today()
    if explicit:
        return explicit
    if mode == "last":
        return _last_week_monday(today)
    if mode == "current":
        return _current_monday(today)
    if mode == "next":
        return _next_monday(today)

    # latest by data
    latest = db.execute(
        select(UserComplexChoice.week_start)
        .distinct()
        .order_by(UserComplexChoice.week_start.desc())
        .limit(1)
    ).scalar_one_or_none()
    return latest or _last_week_monday(today)


@router.get(
    "/choices/last-week.xlsx",
    description=(
        "Экспорт выборов комплексов. По умолчанию — последняя доступная неделя по данным. "
        "Можно указать ?week=last|current|next|latest и/или ?week_start=YYYY-MM-DD"
    )
)
def export_last_week_choices(
    db: Session = Depends(db_session),
    week: str | None = None,
    week_start: date | None = None,
):
    # Determine target week start
    target_week_start = _resolve_week_start(db, week, week_start)

    # Use only weekdays Mon-Fri
    weekdays = [1, 2, 3, 4, 5]

    # Preload available (not closed) complexes per weekday to use as fallback when no choice
    fallback_by_weekday: dict[int, str] = {}
    available_rows = db.execute(
        select(ComplexWeekday.weekday_id, Complex.id, Complex.name)
        .join(Complex, Complex.id == ComplexWeekday.complex_id)
        .where(ComplexWeekday.weekday_id.in_(weekdays), Complex.is_closed == False)
        .order_by(ComplexWeekday.weekday_id, Complex.id)
    ).all()
    for wid, cid, cname in available_rows:
        if wid not in fallback_by_weekday:
            fallback_by_weekday[wid] = cname

    # Fetch users of classes and LEFT JOIN choices for the selected week (only Mon-Fri), exclude class id=1
    stmt = (
        select(
            Class.id.label("class_id"),
            Class.number,
            Class.letter,
            User.id.label("user_id"),
            User.lastname,
            User.name,
            User.patronymic,
            UserComplexChoice.weekday_id,
            Complex.name.label("complex_name"),
        )
        .join(User, User.class_id == Class.id)
        .join(
            UserComplexChoice,
            (UserComplexChoice.user_id == User.id)
            & (UserComplexChoice.week_start == target_week_start)
            & (UserComplexChoice.weekday_id.in_(weekdays)),
            isouter=True,
        )
        .join(Complex, Complex.id == UserComplexChoice.complex_id, isouter=True)
        .where(Class.id != 1)
        .order_by(Class.id, User.lastname, User.name, User.id)
    )

    rows = db.execute(stmt).all()

    # Group rows by class and pivot weekdays to columns per user
    by_class_users: dict[int, dict[int, dict]] = defaultdict(dict)
    class_titles: dict[int, str] = {}

    for r in rows:
        cls_id = r.class_id
        class_titles[cls_id] = f"{r.number or ''}{(r.letter or '').strip()}".strip() or f"class_{cls_id}"
        user_bucket = by_class_users[cls_id].setdefault(
            r.user_id,
            {
                "lastname": r.lastname,
                "name": r.name,
                "patronymic": r.patronymic,
                "choices": {i: "" for i in weekdays},
            },
        )
        if r.weekday_id is not None:
            user_bucket["choices"][int(r.weekday_id)] = r.complex_name or ""

    # Substitute fallbacks for empty choices (Mon-Fri only)
    for users_map in by_class_users.values():
        for u in users_map.values():
            for wid in weekdays:
                if not u["choices"].get(wid):
                    u["choices"][wid] = fallback_by_weekday.get(wid, "")

    # Create workbook with openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover
        raise RuntimeError("openpyxl is required for export. Please add it to requirements and install.")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if not by_class_users:
        ws = wb.create_sheet("No data")
        ws.append(["Нет данных за неделю", str(target_week_start)])
    else:
        weekday_headers = ["Пн", "Вт", "Ср", "Чт", "Пт"]
        for cls_id, users_map in by_class_users.items():
            # Safety check: skip class id 1
            if cls_id == 1:
                continue
            title = class_titles.get(cls_id, f"class_{cls_id}")
            safe_title = (
                title[:31]
                .replace("/", "-")
                .replace("\\", "-")
                .replace("*", "-")
                .replace("[", "(")
                .replace("]", ")")
                .replace(":", "-")
            )
            ws = wb.create_sheet(safe_title or f"class_{cls_id}")

            ws.append([
                "Фамилия",
                "Имя",
                "Отчество",
                *weekday_headers,
            ])

            sorted_users = sorted(
                users_map.items(),
                key=lambda kv: (kv[1]["lastname"] or "", kv[1]["name"] or ""),
            )
            for _, u in sorted_users:
                choices = u["choices"]
                ws.append([
                    u["lastname"],
                    u["name"],
                    u["patronymic"],
                    choices.get(1, ""),
                    choices.get(2, ""),
                    choices.get(3, ""),
                    choices.get(4, ""),
                    choices.get(5, ""),
                ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"choices_{target_week_start.isoformat()}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\""
    }
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
